import os
import openpyxl
from pathlib import Path

lookup = []
required_dbs = []
db_blacklist = [("sys.columns", False, "init"),("sys.indexes", False, "init")]
keyword_db = []


class inputStream:

    def __init__(self, stream) -> None:
        self.stream = stream
        self.index = 0 #points to next unread character
    
    def getEnd(self):
        return len(self.stream)
    
    def peekChar(self):
        if self.index < len(self.stream):
            return self.stream[self.index]
        else:
            return ""
    
    def readChar(self):
        if self.index < len(self.stream):
            self.index=self.index+1
            return self.stream[self.index-1]
        else:
            return ""
    
    def unreadChar(self):
        if self.index != 0:
            self.index = self.index - 1
        return

    def getPosition(self):
        return self.index
    
    def setPosition(self, position):
        self.index = position

def skipSpaces(input):
    c=input.readChar()
    while c == ' ' or c == '\n' or c == '\t' or c =='\v' or c=="\r" or c=="\r\n":
        c=input.readChar()

    input.unreadChar()
    return input

def readString(input, str_start):
    sqlString=[]
    sqlString.append(str_start)

    c = input.readChar()
    while c!=str_start:
        sqlString.append(c)
        c=input.readChar()
    sqlString.append(str_start)

    return "".join(sqlString)

def readComment(input):
    comment = ["-"]

    c=input.readChar()
    while c!= '\n' and c!="":
        comment.append(c)
        c=input.readChar()
    
    return "".join(comment)

def readWord(input):
    word = []
    bracket_stop = True

    c=input.readChar()
    while c!= '\n' and c!=' ' and c!='\t' and c!= "" and c!="\r" and c!="\r\n":
        if c == "$":
            bracket_stop = False
        elif (c=="-" and input.peekChar() == "-") or (c=="*" and input.peekChar() == "/") or c=="\'" or c=="\"" or c==",":
            input.unreadChar()
            break
        elif (c=="(" or c==")") and bracket_stop:
            input.unreadChar()
            bracket_continue = False
            break
        elif (c==")"): #basically one $ only skips one closed bracket; to deal with edge case that a variable ends a bracket
            bracket_stop = True
        
        word.append(c)
        c=input.readChar()

    return "".join(word)

def readBracket(input,endLoc):
    bracket_tokens=[]
    token=""

    input = skipSpaces(input)
    c=input.readChar()
    while c!=")":
        if input.getPosition() == endLoc:
            break
        elif c== '\"' or c== "\'":
            token = readString(input, c)
        elif c == "-" and input.peekChar()=="-":
            token = readComment(input)
        elif c=="/" and input.peekChar()=="*":
            token="/*"
            input.readChar()
        elif c=="*" and input.peekChar() == "/":
            token="*/"
            input.readChar()
        elif c==",":
            token=","
        elif c=="(":
            token=readBracket(input,endLoc)
        else:
            input.unreadChar()
            token=readWord(input)

        if token != "":
            bracket_tokens.append(token)
        
        input = skipSpaces(input)
        c=input.readChar()
    
    return bracket_tokens

def tokenizer(input, tokens):
    token = ""
    
    end_loc=input.getEnd()
    
    input = skipSpaces(input)
    c=input.readChar()
    while c!="":
        if input.getPosition() == end_loc:
            break
        elif c== '\"' or c== "\'":
            token = readString(input, c)
        elif c == "-" and input.peekChar()=="-":
            token = readComment(input)
        elif c=="/" and input.peekChar()=="*":
            token="/*"
            input.readChar()
        elif c==",":
            token=","
        elif c=="*" and input.peekChar() == "/":
            token="*/"
            input.readChar()
        elif c=="(":
            token=readBracket(input,end_loc)
        else:
            input.unreadChar()
            token=readWord(input)
            

        if token != "":
            tokens.append(token)
        
        input = skipSpaces(input)
        c=input.readChar()

    return tokens

def removeExtraChars(string):
    #removes brackets and semicolons from end of db entry
    #mildly deprecated, dont need to remove brackets now
    if string[len(string)-1] == ")" or string[len(string)-1] == ";":
        return string[0:len(string)-1]
    else:
        return string

def dbLookup(token, lookup, db_blacklist):

    #print(lookup)
    #fixes common syntax of just using dbo => workdb.dbo
    if token[0:3].lower() == "dbo":
        token="$(#&USEDB#&)."+token

    #fixes commonly not using dbo at all lol
    elif token[0].casefold() != "$" and token.find("$") == -1 and not searchtf(token, db_blacklist):
        token="$(#&USEDB#&).dbo."+token

    token=removeSqBrackets(token)
    find_val = token.find("$")

    if find_val != -1:
        #replace all variables
        while find_val != -1:
            
            find_val_end = token.find(")")
            search_name=token[find_val+2:find_val_end]
            #print("SEARCHING: "+search_name)
            
            changed = False
            for pair in lookup:
                #print("check4")
                if pair[0].casefold() == search_name.casefold():
                    token=token.replace("$("+search_name+")",(pair[1])[1:len(pair[1])-1])
                    changed = True
            
            if changed == False:
                    token=token.replace("$", "Could not find var = ")

            find_val = token.find("$")
        
        return removeExtraChars(token)

    else: 
        return removeExtraChars(token)

def appendDupCheck(name,array):
    for items in array:
        if items[0].casefold()==name.casefold():
            return False
    
    return True

def reqWordCheck(name,req_words):
    #for filtering words
    for word in req_words:
        if (name.casefold()).find(word.casefold()) == -1:
            return False
    return True

def removeSqBrackets(string):
    string=string.replace("[","")
    string=string.replace("]","")
    return string

def addNewVar(search, replace, lookup, comment_out):
    for index, var_tuple in enumerate(lookup):
        if var_tuple[0].casefold() == search.casefold():
            lookup[index] = (search,replace,comment_out)
            return
        
    lookup.append((search,replace,comment_out))
    return

def searchForDBs(tokens,lookup,required_dbs,db_blacklist,script_name,project, req_words, keyword_db, tokensLen):
    comment_out=False

    for index in range(len(tokens)):
        token=tokens[index]
        if isinstance(token,list):
            searchForDBs(token,lookup,required_dbs,db_blacklist,script_name,project,req_words,keyword_db, tokensLen)
        elif token =="/*":
            comment_out=True
            #print("/*")
        elif token =="*/":
            comment_out=False
            #print("*/")
        elif token.casefold() == ":setvar":
            if comment_out == False:
                addNewVar(tokens[index+1],tokens[index+2], lookup, comment_out)
                index=index+2
                #print("setvar")
        elif token.casefold() == "use":
            pot_working_db = dbLookup(tokens[index+1],lookup,db_blacklist)
            if pot_working_db != tokens[index+1] and comment_out == False: # if dblookup changed something I think?
                if reqWordCheck(pot_working_db,req_words):
                    #print("?? USE " + pot_working_db)
                    addNewVar("#&USEDB#&","\""+pot_working_db+"\"",lookup,comment_out)
        elif token.casefold() == "from":
            if isinstance(tokens[index+1],list):
                keyword_db.append(("from", tokens[index+1], comment_out, script_name, project))
            else:    
                new_entry=dbLookup(tokens[index+1],lookup,db_blacklist)
                if appendDupCheck(new_entry, required_dbs) and reqWordCheck(new_entry,req_words):
                    #print("!!! from: " + new_entry)
                    required_dbs.append((new_entry, comment_out, script_name, project))
                if reqWordCheck(new_entry,req_words):
                    keyword_db.append(("from", new_entry, comment_out, script_name, project))
                index=index+1
        elif token.casefold() == "join":
            if isinstance(tokens[index+1],list):
                keyword_db.append(("join", tokens[index+1], comment_out, script_name, project))
            else:
                new_entry=dbLookup(tokens[index+1],lookup,db_blacklist)
                if appendDupCheck(new_entry, required_dbs) and reqWordCheck(new_entry,req_words):
                    required_dbs.append((new_entry, comment_out, script_name, project))
                if reqWordCheck(new_entry, req_words):
                    keyword_db.append(("join",new_entry, comment_out, script_name, project))
        elif token.casefold() == "into":
            if isinstance(tokens[index+1],list):
                keyword_db.append(("into", tokens[index+1], comment_out, script_name, project))
            else:            
                new_entry=dbLookup(tokens[index+1],lookup,db_blacklist)
                if appendDupCheck(new_entry,db_blacklist):
                    #print("--- into: "+new_entry)
                    db_blacklist.append((new_entry, comment_out, script_name))
                keyword_db.append(("into",new_entry, comment_out, script_name, project))
                index=index+1
        
        elif token.casefold() == "with":
            print(token, end=" || ")
            print(tokens[index+1], end = " || ")
            print(tokens[index+2])
            if not(isinstance(tokens[index+2],list)) and tokens[index+2].casefold() == "as":
                while index+3 < tokensLen and not(isinstance(tokens[index+2],list)) and tokens[index+2].casefold() == "as":
                    print(tokens[index+1], end= " ")
                    print(tokens[index+2], end= " ")
                    print(tokens[index+3])
                    if appendDupCheck(tokens[index+1],db_blacklist):
                        db_blacklist.append((removeSqBrackets(tokens[index+1]), comment_out, script_name))
                    keyword_db.append(("with",tokens[index+1], comment_out, script_name, project))
                    if isinstance(tokens[index+3], list):
                        searchForDBs(tokens[index+3],lookup,required_dbs,db_blacklist,script_name,project,req_words,keyword_db, tokensLen)
                    index=index+4
                    #print("with")
        elif token.casefold() == "drop":
            if isinstance(tokens[index+1],list):
                continue
            else:
                if tokens[index+1].casefold() == "view" or tokens[index+1].casefold() == "table":
                    if isinstance(tokens[index+2],list) or isinstance(tokens[index+3],list):
                        continue
                    else:
                        if tokens[index+2].casefold() == "if" and tokens[index+3].casefold() == "exists":
                            new_entry=dbLookup(tokens[index+4],lookup,db_blacklist)
                            index=index+2
                        else:
                            new_entry=dbLookup(tokens[index+2],lookup,db_blacklist)
                        if appendDupCheck(new_entry,db_blacklist):
                            #print("+++ drop: "+new_entry)
                            db_blacklist.append((new_entry, comment_out, script_name))
                            index=index+1
                        keyword_db.append(("drop", new_entry, comment_out, script_name, project))
                    index=index+1
        elif token.casefold() == "update":
            if isinstance(tokens[index+1],list):
                keyword_db.append(("update", tokens[index+1], comment_out, script_name, project))
            else:            
                new_entry = dbLookup(tokens[index+1],lookup,db_blacklist)
                if appendDupCheck(new_entry,db_blacklist):
                    db_blacklist.append((new_entry, comment_out, script_name))
                index=index+1
                keyword_db.append(("update", new_entry, comment_out, script_name, project))

def searchtf(string, array):
    for items in array:
        if items[0].casefold()==string.casefold():
            return True
    return False

def remove_bl_dbs(required_dbs,db_blacklist):
    pop_offset=0
    pop_queue = []

    for k in range(len(required_dbs)):    
        if searchtf((required_dbs[k])[0],db_blacklist):
           pop_queue.append(k)

        #HAVE TO FIND A STRING WHICH IS NEEDED FOR ANY DB ENTRY HERE, I guess by user input?      
        #elif len((required_dbs[k])[0]) < len("BASHAS") or (required_dbs[k])[0].upper().find("BASHAS") == -1:
        #    pop_queue.append(k)
        #UPD: MOVED UP APPENDDUPCHECK FOR EFFICIENCY

    #print(pop_queue)

    for to_pop in pop_queue:
        required_dbs.pop(to_pop-pop_offset)
        pop_offset=pop_offset+1

def scriptInterper(file_path, lookup, required_dbs, db_blacklist, project, req_words, keyword_db):
    cont=True
    tokens=[]
    fd=open(file_path)
    inpStream = inputStream(fd.read())
    script_name = os.path.split(file_path)[1]
    #print("\nSCRIPT interper starting on... "+script_name)

    tokenizer(inpStream, tokens)
    searchForDBs(tokens,lookup,required_dbs,db_blacklist,script_name, project, req_words, keyword_db, len(tokens))
    remove_bl_dbs(required_dbs,db_blacklist)
    
    #print("\n\n\n")
    #for token in tokens:
    #    print(token, end=' || ')
    #print("\n")
    #print(len(tokens))

    cont=True

def folderInterper(folder_path, lookup, required_dbs, db_blacklist, req_words, keyword_db):
    script_queue = []
    project=os.path.split(folder_path)[1]
    
    folder_contents = os.listdir(folder_path) 
    
    for file in folder_contents:
#        print(os.path.join(folder_path, file))
#        print(os.path.splitext(os.path.join(folder_path, file))[1].casefold())
        if os.path.splitext(os.path.join(folder_path, file))[1].casefold() == ".sql".casefold():
            script_queue.append(os.path.join(folder_path,file))
    
    for script in script_queue:
        print("\n"+script)
        scriptInterper(script, lookup, required_dbs, db_blacklist,project,req_words, keyword_db)
    

    #os.path.splitext(filepath)[1]

def getNumber(unsorted_item):
    return int((unsorted_item[0])[0])

def projectInterper(project_path,lookup,required_dbs,db_blacklist,req_words,keyword_db):
    folder_contents = os.listdir(project_path)
    folder_queue = []

    for item in folder_contents:
        sub_path = os.path.join(project_path,item)
        if os.path.isdir(sub_path):
            folder_queue.append(((os.path.split(sub_path)[1]).split("_"),sub_path))

    folder_queue.sort(key=getNumber)

    #print(folder_queue)

    for queue_item in folder_queue:
        folder=queue_item[1]
        print(folder)
        folderInterper(folder,lookup,required_dbs,db_blacklist,req_words,keyword_db)


projectInterper(r"C:\Users\rober\Documents\Coding\Work_Projects\PRGX\2020", lookup, required_dbs, db_blacklist, [], keyword_db)

#folderInterper(r"C:\Users\rober\Documents\Coding\Work_Projects\PRGX\2020\1_AP", lookup, required_dbs, db_blacklist,["Bashas"], keyword_db)

#scriptInterper(r"C:\Users\rober\Documents\Coding\Work_Projects\PRGX\2020\8_MVMT\02_BAS_WCSGGD_SETUP.sql", lookup, required_dbs,db_blacklist,"AP",[], keyword_db)

# testpath=Path(r"C:\Users\rober\Documents\Coding\Work_Projects\PRGX\BASHAS 1_AP")
# lst=os.listdir(testpath)
# print(lst)
# print(type(os.path.splitext(os.path.join(testpath, lst[1]))[1]))

# tokenizer(tokens)
# searchForDBs(tokens,lookup,required_dbs)
# remove_bl_dbs(required_dbs,db_blacklist)
print("\n\n\n")
print(required_dbs)
#print("\n\n\n")
#print(db_blacklist)
#print("\n\n\n")
#print(lookup)

def listToString(lst):
    output="("
    for index,entry in enumerate(lst):
        if isinstance(entry, list):
            output= output+listToString(entry)
        else:
            output=output+entry+" "
    if len(output)>1:
        output=output[:len(output)-1] +")"
    else: output=output+")"
    return output

workbook = openpyxl.Workbook()

reqsheet= workbook.create_sheet("db required",0)
for index, entry in enumerate(required_dbs):
    cellref=reqsheet.cell(row=index+2, column=3)
    cellref.value=entry[0]

    cellref2=reqsheet.cell(row=index+2, column=4)
    if entry[1] == True:
        cellref2.value="commented out,"
    if entry[0].casefold().find("_w.".casefold()) != -1:
        cellref2.value="potential workingDB,"+ str(cellref2.value or "")
    if isinstance(cellref2.value, str):
        cellref2.value = cellref2.value[:-1]

    cellref3=reqsheet.cell(row=index+2, column=2)
    cellref3.value=entry[2]
    cellref4=reqsheet.cell(row=index+2, column=1)
    cellref4.value=entry[3]
(reqsheet.cell(row=1,column=1)).value="Project"
(reqsheet.cell(row=1,column=2)).value="Script Name"
(reqsheet.cell(row=1,column=3)).value="Database Name"
(reqsheet.cell(row=1,column=4)).value="Extra info"
#reqsheet.column_dimensions['A'].width = 20
reqsheet.column_dimensions['B'].width = 30
reqsheet.column_dimensions['C'].width = 50

blsheet=workbook.create_sheet("black list",1)
for index, entry in enumerate(db_blacklist):
    cellref=blsheet.cell(row=index+2,column=2)
    cellref.value=entry[0]
    if entry[1] == True:
        cellref=blsheet.cell(row=index+2, column=3)
        cellref.value="commented out"
    #script name
    cellref4=blsheet.cell(row=index+2, column=1)
    cellref4.value=entry[2]
(blsheet.cell(row=1,column=1)).value="Script Name"
(blsheet.cell(row=1,column=2)).value="Database Name"
(blsheet.cell(row=1,column=3)).value="Extra info"

setvars=workbook.create_sheet("setvars",2)
for index, entry in enumerate(lookup):
    cellref=setvars.cell(row=index+1,column=1)
    cellref2=setvars.cell(row=index+1,column=2)
    cellref.value=entry[0]
    cellref2.value=entry[1]
    if entry[2] == True:
        cellref=setvars.cell(row=index+1, column=3)
        cellref.value="commented out"

keywords=workbook.create_sheet("full list",3)
(keywords.cell(row=1,column=1)).value="Keyword"
(keywords.cell(row=1,column=2)).value="Database/Input"
(keywords.cell(row=1,column=3)).value="Comment out T/F"
(keywords.cell(row=1,column=4)).value="Script"
(keywords.cell(row=1,column=5)).value="Project"
keywords.column_dimensions['B'].width = 50
keywords.column_dimensions['D'].width = 30
#keywords.column_dimensions['E'].width = 20
for index, entry in enumerate(keyword_db):
    for i in range(5):
        cellref=keywords.cell(row=index+2,column=i+1)
        if isinstance(entry[i],list):
            
            cellref.value = listToString(entry[i]) #' '.join(entry[i])
        else:
            cellref.value=entry[i]


workbook.save(filename=r'C:\Users\rober\Documents\Coding\Work_Projects\PRGX\SQLInterper\sqlinterperROUGHResults.xlsx')

